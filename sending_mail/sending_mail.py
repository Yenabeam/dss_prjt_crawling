#-----------------------메일신청자불러오기---------------------
import pymysql
import pandas as pd 

def load_mail():
    conn = pymysql.connect(host = '00.000.00.0', user = 'user', password = 'user' ,db = 'UserInfo')
    curs = conn.cursor()

    sql = "SELECT * FROM UserInfo"
    curs.execute(sql)
    rows = curs.fetchall()
    conn.close()

    list = []
    for i in rows:
        dic = {
            'user' : i[0],
            'price' : i[1],
            'email' : i[2]
        }
        list.append(dic)
        df = pd.DataFrame(list)
        df.to_excel('subscribe.xlsx',encoding='utf-8')
    return df 
load_mail()
#-----------------------메일발송함수---------------------

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
# SMTP 접속을 위한 서버, 계정 설정
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465

# 보내는 메일 계정
SMTP_USER = "USER@gmail.com"
SMTP_PASSWORD = 'PASSWORD'


def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False
    

# 이메일 보내기 함수
def send_mail(addr, subj_layout, cont_layout, attachment=None):
    if not is_valid(addr):
        print("Wrong email: " + addr)
        return
            
    # 텍스트 파일
    msg = MIMEMultipart("alternative")
    # 첨부파일이 있는 경우 mixed로 multipart 생성
    if attachment:
        msg = MIMEMultipart('mixed')
    msg["From"] = SMTP_USER
    msg["To"] = addr
    msg["Subject"] = subj_layout
    contents = cont_layout
    text = MIMEText(_text = contents, _charset = "utf-8")
    msg.attach(text)

    # smtp로 접속할 서버 정보를 가진 클래스변수 생성
    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    # 해당 서버로 로그인
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    # 메일 발송
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    # 닫기
    smtp.close()

#-----------------------메일보내기---------------------
from openpyxl import load_workbook
wb = load_workbook('subscribe.xlsx')
ws = wb.active
for row in ws.iter_rows():
    addr = row[1].value
    subj_layout = '[fleafully]{}님에게 누구보다 빠르게 최저가알림'.format(row[3].value)
    cont_layout = '요청하신 {}원 이하의 물건 정보 입니다.'.format(row[2].value)
    send_mail(addr, subj_layout, cont_layout)